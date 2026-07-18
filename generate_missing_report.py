import openpyxl
from collections import defaultdict
import os

# File paths
flat_path = '/Users/akshatojha/Downloads/sail data.xlsx'
dist_path = '/Users/akshatojha/Downloads/SAIL/PC PRINTER MFD SCANNER DETAILS 17 04 2026 - Copy.xlsx'
report_path = '/Users/akshatojha/.gemini/antigravity-cli/brain/09e56e38-970b-40f0-8d2b-525c2ce7a5b4/sail_missing_tags_report.md'

print("Starting audit comparison...")

# Load flat database
wb_flat = openpyxl.load_workbook(flat_path, data_only=True)
ws_flat = wb_flat.active
flat_rows = list(ws_flat.iter_rows(values_only=True))
flat_headers = flat_rows[0]
flat_clean = [r for r in flat_rows[1:] if any(val is not None and str(val).strip() != '' for val in r)]

# Load distribution sheets
wb_dist = openpyxl.load_workbook(dist_path, data_only=True)

# 1. Parse Tagged Counts
flat_dept_idx = flat_headers.index('Deptt.')
pc_make_idx = flat_headers.index('PC Make')
printer_make_idx = flat_headers.index('Printer Make ')
mfd_make_idx = flat_headers.index('MFD MAKE')
scanner_make_idx = flat_headers.index('Scanner  Make ')

tagged_by_dept = defaultdict(lambda: {'PC': 0, 'Printer': 0, 'MFD': 0, 'Scanner': 0})
total_tagged = {'PC': 0, 'Printer': 0, 'MFD': 0, 'Scanner': 0}

for r in flat_clean:
    dept = str(r[flat_dept_idx]).strip().upper() if r[flat_dept_idx] is not None else 'UNKNOWN'
    if dept.startswith('CGM'):
        dept = 'CGM ELECTRICAL MAINTENANCE'
        
    if r[pc_make_idx] is not None and str(r[pc_make_idx]).strip() != '':
        tagged_by_dept[dept]['PC'] += 1
        total_tagged['PC'] += 1
    if r[printer_make_idx] is not None and str(r[printer_make_idx]).strip() != '':
        tagged_by_dept[dept]['Printer'] += 1
        total_tagged['Printer'] += 1
    if r[mfd_make_idx] is not None and str(r[mfd_make_idx]).strip() != '':
        tagged_by_dept[dept]['MFD'] += 1
        total_tagged['MFD'] += 1
    if r[scanner_make_idx] is not None and str(r[scanner_make_idx]).strip() != '':
        tagged_by_dept[dept]['Scanner'] += 1
        total_tagged['Scanner'] += 1

# 2. Parse Allotted Counts
allotted_by_dept = defaultdict(lambda: {'PC': 0, 'Printer': 0, 'MFD': 0, 'Scanner': 0})
total_allotted = {'PC': 0, 'Printer': 0, 'MFD': 0, 'Scanner': 0}

def add_allotted(sheet_name, device_type, total_col_label, header_row_idx):
    ws = wb_dist[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[header_row_idx]
    
    total_idx = -1
    for c_idx, val in enumerate(headers):
        if val is not None and total_col_label.upper() in str(val).upper():
            total_idx = c_idx
            break
    if total_idx == -1:
        total_idx = len(headers) - 1
        
    for r in rows[header_row_idx+1:]:
        dept = str(r[1]).strip().upper() if r[1] is not None else ''
        if dept and dept not in ['TOTAL', 'GRAND TOTAL', 'DEPTT.', 'None']:
            if dept.startswith('CGM'):
                dept = 'CGM ELECTRICAL MAINTENANCE'
            val = r[total_idx]
            if val is not None:
                try:
                    qty = int(val)
                    allotted_by_dept[dept][device_type] += qty
                    total_allotted[device_type] += qty
                except ValueError:
                    pass

add_allotted('PC DETAILS', 'PC', 'Grand Total', 1)
add_allotted('PRINTER DETAIL', 'Printer', 'Grand Total', 1)
add_allotted('MFD DETAIL', 'MFD', 'Grand Total', 1)
add_allotted('SCANNER DETAILS', 'Scanner', 'TOTAL', 2)

# Generate Markdown Report
markdown = []
markdown.append("# SAIL Asset Management Tagging Discrepancy Audit Report\n")
markdown.append("This report compares the official equipment quantities allotted to departments against the actual tagged items recorded in the asset tracking database.\n")

# Summary Table
markdown.append("## 1. Executive Summary")
markdown.append("| Equipment Type | Total Allotted (Distribution) | Total Tagged (Database) | Missing Tags (Deficit) |")
markdown.append("| :--- | :---: | :---: | :---: |")

for dtype in ['PC', 'Printer', 'MFD', 'Scanner']:
    allot = total_allotted[dtype]
    tag = total_tagged[dtype]
    diff = allot - tag
    markdown.append(f"| {dtype} | {allot} | {tag} | **{diff}** |")

tot_allot = sum(total_allotted.values())
tot_tag = sum(total_tagged.values())
tot_diff = tot_allot - tot_tag
markdown.append(f"| **TOTAL** | **{tot_allot}** | **{tot_tag}** | **{tot_diff}** |\n")

# Department Table
markdown.append("## 2. Department-Wise Deficit Breakdown")
markdown.append("The table below lists departments that have a positive deficit (allotted equipment that is not yet tagged in the database).\n")
markdown.append("| Department Name | PC (Allot/Tag) | Printer (Allot/Tag) | MFD (Allot/Tag) | Scanner (Allot/Tag) | Total Missing Tags |")
markdown.append("| :--- | :---: | :---: | :---: | :---: | :---: |")

all_depts = set(tagged_by_dept.keys()).union(set(allotted_by_dept.keys()))
sorted_depts = sorted(list(all_depts))

for dept in sorted_depts:
    if not dept.strip() or dept == 'UNKNOWN':
        continue
    allot = allotted_by_dept[dept]
    tag = tagged_by_dept[dept]
    
    pc_diff = allot['PC'] - tag['PC']
    prn_diff = allot['Printer'] - tag['Printer']
    mfd_diff = allot['MFD'] - tag['MFD']
    scn_diff = allot['Scanner'] - tag['Scanner']
    
    total_deficit = pc_diff + prn_diff + mfd_diff + scn_diff
    
    if total_deficit > 0:
        markdown.append(
            f"| {dept} | "
            f"{allot['PC']}/{tag['PC']} ({pc_diff}) | "
            f"{allot['Printer']}/{tag['Printer']} ({prn_diff}) | "
            f"{allot['MFD']}/{tag['MFD']} ({mfd_diff}) | "
            f"{allot['Scanner']}/{tag['Scanner']} ({scn_diff}) | "
            f"**{total_deficit}** |"
        )

# Write report to folder
os.makedirs(os.path.dirname(report_path), exist_ok=True)
with open(report_path, 'w') as f:
    f.write("\n".join(markdown))
print(f"Audit report generated successfully at: {report_path}")
